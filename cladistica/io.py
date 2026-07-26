from __future__ import annotations

import csv
from collections import OrderedDict
from pathlib import Path


def read_tsv(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    with path.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        rows = [{key: value or "" for key, value in row.items()} for row in reader]
        return rows, list(reader.fieldnames or [])


def read_csv(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = [{key: value or "" for key, value in row.items()} for row in reader]
        return rows, list(reader.fieldnames or [])


def read_xlsx(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read .xlsx files. Install dependencies from requirements.txt.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    header_values = next(rows_iter, None)
    if not header_values:
        return [], []
    fieldnames = [str(value).strip() if value is not None else "" for value in header_values]
    rows: list[dict[str, str]] = []
    for values in rows_iter:
        row = {
            fieldnames[index]: "" if value is None else str(value)
            for index, value in enumerate(values)
            if index < len(fieldnames) and fieldnames[index]
        }
        if any(value for value in row.values()):
            rows.append(row)
    return rows, [field for field in fieldnames if field]


def read_table(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return read_xlsx(path)
    if suffix == ".csv":
        return read_csv(path)
    return read_tsv(path)


def write_tsv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: "" if row.get(key) is None else str(row.get(key, "")) for key in fieldnames})


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: "" if row.get(key) is None else str(row.get(key, "")) for key in fieldnames})


def write_xlsx(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> bool:
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError:
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "accession_summary"
    sheet.append(fieldnames)
    for row in rows:
        sheet.append(["" if row.get(key) is None else row.get(key, "") for key in fieldnames])
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 48)
    workbook.save(path)
    return True


def read_fasta(path: Path) -> OrderedDict[str, str]:
    records: OrderedDict[str, str] = OrderedDict()
    header = ""
    chunks: list[str] = []
    if not path.exists():
        return records
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line:
            continue
        if line.startswith(">"):
            if header:
                records[header] = "".join(chunks)
            header = line[1:].strip()
            chunks = []
        else:
            chunks.append(line)
    if header:
        records[header] = "".join(chunks)
    return records


def write_fasta(path: Path, records: dict[str, str], width: int = 80) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []
    for header, sequence in records.items():
        lines.append(f">{header}")
        seq = sequence.strip()
        for start in range(0, len(seq), width):
            lines.append(seq[start : start + width])
    path.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8")


def sample_id_from_header(header: str) -> str:
    return header.split("|", 1)[0].strip()


def find_first_existing(paths: list[Path]) -> Path | None:
    return next((path for path in paths if path.exists()), None)
